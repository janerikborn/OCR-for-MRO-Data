# Modul 1: Import von Bibliotheken und Definition der Pfade

import time  # Zeitmessung der Modulausführung
import cv2  # OpenCV-Bibliothek für Bildverarbeitung
import numpy as np  # Für mathematische Operationen mit Arrays
import random  # Zum Generieren von Zufallszahlen
import os  # Um mit dem Dateisystem zu arbeiten
from pdf2image import convert_from_path, pdfinfo_from_path  # Um PDF-Seiten in Bilder umzuwandeln
from concurrent.futures import ThreadPoolExecutor  # Ermöglicht parallele Verarbeitung
import gc  # Für die Speicherbereinigung
import pandas as pd # Datenmanipulation und -analyse, Organisation der extrahierten Daten
from openpyxl import load_workbook # Lesen und Schreiben von Excel-Dateien zur Speicherung der Ergebnisse
from openpyxl.styles import Font # Schreiben von Excel-Dateien

import pytesseract  # Open Source Text-OCR Software
from google.cloud import vision # Cloudbasierte Texterkennung

# Zentrale Definition der Input- und Output-Verzeichnisse

# Hauptverzeichnis für die 1. PDF
base_dir = 'C:/Users/Katharina/OCR - PDF_to_Excel/1. PDF Galistair'

# Pfade für Eingabe- und Ausgabedateien
pdf_file = os.path.join(base_dir, '0876_OCCM_Status_GTR_as_of_22MAR24.pdf')
excel_comparison_file = os.path.join(base_dir, 'Referenzdaten_0876_OCCM Status_GTR as of 22MAR24.xlsx')

# Verzeichnisse für Zwischenergebnisse und Ausgaben
input_dir_png = os.path.join(base_dir, '1_PDF_to_PNG')
output_dir_png = os.path.join(base_dir, '1_PDF_to_PNG')
input_dir_roi = os.path.join(base_dir, '2_PNG_to_ROI')
output_dir_roi = os.path.join(base_dir, '2_PNG_to_ROI')
output_dir_results = os.path.join(base_dir, '3_Ergebnisse')

# Pfade zu spezifischen Dateien
tesseract_cmd = r'C:\Users\Katharina\AppData\Local\Programs\Tesseract-OCR\tesseract.exe' # Pfad zur Lokal gespeicherten Pytesseract Datei - Installationstutorial: https://www.youtube.com/watch?v=O8maBz1yXe0 Downloadlink (für Windows): https://github.com/UB-Mannheim/tesseract/wiki 
google_vision_credentials = 'C:/Users/Katharina/OCR - PDF_to_Excel/vision-api.json' # Pfad zur lokal gespeichterten Vision API .json Datei - Installationstutorial: https://www.youtube.com/watch?v=OFrwjBbk9SY bis Minute 2:10
poppler_path = r"C:\Users\Katharina\poppler-24.07.0\Library\bin"  # Pfad zum Poppler \Library\bin Ordner - Download/Installtion: https://www.geeksforgeeks.org/convert-pdf-to-image-using-python/

# Sicherstellen, dass die Output-Verzeichnisse existieren
os.makedirs(output_dir_png, exist_ok=True)
os.makedirs(output_dir_roi, exist_ok=True)
os.makedirs(output_dir_results, exist_ok=True)


# Modul 2: Jede PDF-Seite in eine PNG-Datei umwandeln und speichern (mit Hilfe von PDF2Image und Poppler)

start_time = time.time()

# Diese Funktion nimmt eine PDF-Datei und wandelt jede Seite in ein hochauflösendes PNG-Bild mit 350 dpi (Dots Per Inch") um
def extract_pdf_to_png(pdf_file, output_folder, dpi=350):
    # Sicherstellen, dass der Ausgabeordner existiert
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Gesamtanzahl der Seiten ermitteln
    info = pdfinfo_from_path(
        pdf_file,
        poppler_path=poppler_path  # Verwendung des zentral definierten Poppler-Pfades
    )
    max_pages = info["Pages"]

    # Seiten einzeln verarbeiten, um Speicherüberlastung zu vermeiden
    for page_number in range(1, max_pages + 1):
        images = convert_from_path(
            pdf_file,
            dpi=dpi,
            first_page=page_number,
            last_page=page_number,
            poppler_path=poppler_path  # Verwendung des zentral definierten Poppler-Pfades
        )
        for image in images:
            output_path = os.path.join(output_folder, f'page_{page_number}.png')
            image.save(output_path, 'PNG')
            print(f"Seite {page_number} gespeichert unter {output_path}")

            # Speicher für jede Seite explizit freigeben
            del image
            gc.collect()  # Speicherbereinigung durchführen
        # Auch den Speicher für die gesamte Liste `images` freigeben
        del images
        gc.collect()

    print(f"Alle Seiten wurden erfolgreich in {output_folder} gespeichert.")

# Aufruf der Funktion mit den zentral definierten Pfaden
extract_pdf_to_png(pdf_file, output_dir_png)

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")


# Modul 3: Erkennung der ROI's ("bounding boxes") innerhalb horizontaler Linien, welche alle Daten eines repariertes Flugzeugteils enthalten.

# Funktion zur Verarbeitung von Bildern und Extraktion von ROIs
def process_and_extract_rois(input_dir, output_dir, min_width=100, min_height=20, line_threshold=70):
# - min_width: Minimale Breite eines ROIs (100 Pixel).
# - min_height: Minimale Höhe eines ROIs (20 Pixel).
# - line_threshold: Minimale vertikale Distanz zwischen horizontalen Linien, um einen neuen ROI zu erzeugen (70 Pixel).

    # Ausgabeordner erstellen, falls nicht vorhanden
    os.makedirs(output_dir, exist_ok=True)
    i = 0  # Zähler für die gespeicherten ROIs

    # Bilddateien im Eingabeordner sortiert sammeln
    image_files = sorted(
        [f for f in os.listdir(input_dir) if f.endswith('.png')],
        key=lambda x: int(x.split('_')[1].split('.')[0])
    )
    print("Zu verarbeitende Dateien:", image_files)

    for filename in image_files:
        image_path = os.path.join(input_dir, filename)
        image = cv2.imread(image_path) # Bild mit OpenCV laden


        # Prüfen, ob das Bild geladen wurde
        if image is None:
            print(f"Bild konnte nicht geladen werden: {image_path}")
            continue

        # Das Bild in Graustufen umwandeln, um die Kanten einfacher zu erkennen
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Kanten im Graustufenbild erkennen, um Linien hervorzuheben
        edges = cv2.Canny(gray, 50, 150, apertureSize=3)

        # Erstellen eines 1-Pixel-hohen, 50-Pixel-breiten Filters (Kernel), um nur horizontale Linien zu extrahieren
        kernel = np.ones((1, 50), np.uint8)
        horizontal_lines = cv2.morphologyEx(edges, cv2.MORPH_OPEN, kernel)

        # Finden der Konturen der horizontalen Linien im Bild
        contours, _ = cv2.findContours(horizontal_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        # Mindestbreite einer Linie berechnen: 90% der Bildbreite, um nur lange Linien zu erkennen
        image_width = image.shape[1]
        min_contour_width = int(0.9 * image_width)
        # Filtern der Konturen, um nur die ausreichend langen Linien zu behalten
        long_contours = [c for c in contours if cv2.boundingRect(c)[2] >= min_contour_width]

        # Konturen nach ihrer y-Position sortieren (von oben nach unten)
        long_contours = sorted(long_contours, key=lambda c: cv2.boundingRect(c)[1])

        # Liste zum Speichern der Begrenzungsrahmen (ROIs) zwischen den horizontalen Linien
        bounding_boxes = []
        # Anzahl der horizontalen Linien im Bild
        num_long_lines = len(long_contours)

        # Schleife, um Begrenzungsrahmen (ROIs) zwischen den erkannten horizontalen Linien zu erstellen
        for j in range(3, num_long_lines - 1):  # Startet bei der 4. Linie (Index 3)
            x, y, w, h = cv2.boundingRect(long_contours[j])
            next_line_y = cv2.boundingRect(long_contours[j + 1])[1]
            # Prüfen, ob der Abstand zur nächsten Linie größer als der Schwellenwert ist
            if next_line_y - y > line_threshold:
                bounding_boxes.append((0, y, image.shape[1], next_line_y))

        # Bildabmessungen erhalten
        height, width = image.shape[:2]

        # Schleife zum Durchlaufen aller erkannten Begrenzungsrahmen (ROIs) und Extraktion der Bereiche
        for idx, (x1, y1, x2, y2) in enumerate(bounding_boxes):
            # Überprüfen, ob die Koordinaten des Begrenzungsrahmens gültig sind
            x1, x2 = max(0, x1), min(width, x2)
            y1, y2 = max(0, y1), min(height, y2)

            # Dimensionen berechnen
            box_width = x2 - x1
            box_height = y2 - y1

            # Überspringen, wenn zu klein
            if box_width < min_width or box_height < min_height:
                continue

            # Ausschneiden des ROIs aus dem Bild
            roi = image[y1:y2, x1:x2]
            # Erstellen eines eindeutigen Dateinamens für den ROI und Speichern im Ausgabeordner
            roi_filename = os.path.join(output_dir, f'roi_{i + 1}.png')
            cv2.imwrite(roi_filename, roi)
            i += 1  # Zähler erhöhen

    print("Verarbeitung abgeschlossen. ROIs gespeichert in", output_dir)

start_time = time.time()

# Aufruf der Funktion mit den zentral definierten Pfaden
process_and_extract_rois(input_dir_png, output_dir_roi)

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")


# Modul 4: Formel, um Text mit Tesseract aus einem ROI("Begrenzungsrahmen")-Bild zu extrahieren

# Pfad zur Tesseract-OCR-Programmdatei festlegen
pytesseract.pytesseract.tesseract_cmd = tesseract_cmd  # Verwendung des zentral definierten Pfades

# Funktion zur Textextraktion aus einem Bild
# - x und y: Startkoordinaten für die Textextraktionsbereiche im ROI-Bild
def extract_text(image_path, x=0, y=0):
    # Bild laden
    image = cv2.imread(image_path)
    width = 360
    height = 80
    # Die Breite und Höhe (360 und 80 Pixel) werden hier festgelegt, um 
    # nur einen bestimmten Bildbereich für die Texterkennung zu betrachten.

    # Prüfen, ob das Bild geladen wurde
    if image is None:
        raise ValueError(f"Bild konnte nicht geladen werden. Überprüfe den Pfad: {image_path}")

    # Definiere den Bereich für die Textextraktion (Textextraktionsbereiche)
    roi = image[y:y+height, x:x+width]

    # Überprüfen, ob der ausgewählte Bereich gültig ist
    if roi is None or roi.size == 0:
        raise ValueError(f"ROI ist leer. Überprüfe die Abmessungen: Breite={width}, Höhe={height}")

    # Den Textextraktionsbereich in Graustufen konvertieren 
    roi_gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)

    # Texterkennung im Textextraktionsbereich mit Tesseract als Zeichenkette
    text = pytesseract.image_to_string(roi_gray)

    print(f"Verarbeitung abgeschlossen für: {image_path}")
    # Entfernt überflüssige Leerzeichen vom Anfang und Ende des Textes
    return text.strip()  


# Modul 5: Funktionsaufruf von Textextraktion mit Pytesseract und Speicherung in einer Excel-Datei - (1. Ergebnis ohne Eliminierug von häufigsten Fehlern)

start_time = time.time()

def extract_texts_to_excel(input_dir, output_file):
    ataChapter = []  # Liste für die gesammelten "ATA-Kapitel"
    partNumber = []  # Liste für die gesammelten "Teilenummern"
    serialNumber = []  # Liste für die gesammelten "Seriennummern"

    # Bilddateien im Eingabeordner sortiert sammeln
    image_files = sorted([f for f in os.listdir(input_dir) if f.endswith('.png')],
                         key=lambda x: int(x.split('_')[1].split('.')[0]))

    def process_image(filename):
        image_path = os.path.join(input_dir, filename)

        # Extrahiere Text aus drei Textextrkationbereichen des Bildes
        t1 = extract_text(image_path)  # Standardbereich (x=0, y=0), enthält ataChapter
        t2 = extract_text(image_path, x=1120)  # Bereich ab x=1120, y=0, enthält partNumber
        t3 = extract_text(image_path, x=1585)  # Bereich ab x=1585, y=0, enthält serialNumber
        return t1, t2, t3

    # ThreadPoolExecutor nutzen, um Bilder parallel zu verarbeiten
    with ThreadPoolExecutor() as executor:
        results = executor.map(process_image, image_files)

    # Ergebnisse sammeln
    for t1, t2, t3 in results:
        ataChapter.append(t1)
        partNumber.append(t2)
        serialNumber.append(t3)

    # Daten in ein DataFrame umwandeln
    data = {
        'ataChapter': ataChapter,
        'partNumber': partNumber,
        'serialNumber': serialNumber
    }
    df = pd.DataFrame(data)

    # DataFrame in Excel speichern
    df.to_excel(output_file, index=False)
    print(f"Daten erfolgreich in Excel gespeichert unter {output_file}.")

# Aufruf der Funktion mit den zentral definierten Pfaden
extract_texts_to_excel(output_dir_roi, os.path.join(output_dir_results, '1.Ergebnis_mit_Pytesseract.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")


# Modul 6: Funktionsaufruf von Textextraktion mit Pytesseract und Speicherung in einer 2. Excel-Datei - 2. Ergebnis mit Eliminierug der häufigsten Fehler (hartkodiert)

start_time = time.time()

def extract_texts_to_excel_cleaned(input_dir, output_file):
    ataChapter = []
    partNumber = []
    serialNumber = []
    image_files = sorted([f for f in os.listdir(input_dir) if f.endswith('.png')],
                         key=lambda x: int(x.split('_')[1].split('.')[0]))

    def process_image(filename):
        image_path = os.path.join(input_dir, filename)
        # Häufige Fehler eliminieren, ersetzt "$" durch "S" und Leerzeichen entfernt
        t1 = extract_text(image_path).replace(" ", "").replace("$", "S")
        t2 = extract_text(image_path, x=1120).replace(" ", "").replace("$", "S")
        t3 = extract_text(image_path, x=1585).replace(" ", "").replace("$", "S")
        return t1, t2, t3

    # Bilder parallel verarbeiten
    with ThreadPoolExecutor() as executor:
        results = executor.map(process_image, image_files)

    # Ergebnisse sammeln
    for t1, t2, t3 in results:
        ataChapter.append(t1)
        partNumber.append(t2)
        serialNumber.append(t3)

    # Daten in ein DataFrame umwandeln
    data = {
        'ataChapter': ataChapter,
        'partNumber': partNumber,
        'serialNumber': serialNumber
    }
    df = pd.DataFrame(data)

    # DataFrame in Excel speichern
    df.to_excel(output_file, index=False)
    print(f"Daten erfolgreich in Excel gespeichert unter {output_file}.")

# Aufruf der Funktion mit den zentral definierten Pfaden
extract_texts_to_excel_cleaned(output_dir_roi, os.path.join(output_dir_results, '2.Ergebnis_mit_Pytesseract.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")


# Modul 7: Formel, um Text mit Google Vision API aus einem ROI("Begrenzungsrahmen")-Bild zu extrahieren

# Pfad zur Google Vision API festlegen
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = google_vision_credentials  # Verwendung des zentral definierten Pfades

client = vision.ImageAnnotatorClient()

def extract_text_google_vision(image_path, x=0, y=0):
    # Bild laden
    time.sleep(1)
    image = cv2.imread(image_path)
    width = 360
    height = 80
    # Prüfen, ob das Bild erfolgreich geladen wurde
    if image is None:
        raise ValueError(f"Bild konnte nicht geladen werden. Überprüfe den Pfad: {image_path}")

    # ROI (Region of Interest) definieren
    roi = image[y:y+height, x:x+width]

    # Überprüfen, ob der ROI gültig ist
    if roi is None or roi.size == 0:
        raise ValueError(f"ROI ist leer. Überprüfe die Abmessungen: Breite={width}, Höhe={height}")

    # Bild für die Google Vision API vorbereiten und Text extrahieren
    success, encoded_image = cv2.imencode('.png', roi)
    content = encoded_image.tobytes()
    image = vision.Image(content=content)
    response = client.text_detection(image=image)
    texts = response.text_annotations
    full_text = []
    for text in texts:
        full_text.append(text.description)
    print(f"Verarbeitung abgeschlossen für: {image_path}")
    return full_text


# Modul 8: Funktionsaufruf von Textextraktion mit Google Vision API und Speicherung in einer 3. Excel-Datei 

start_time = time.time()

def extract_texts_to_excel_google_vision(input_dir, output_file):
    ataChapter = []
    partNumber = []
    serialNumber = []

    image_files = sorted(
        [f for f in os.listdir(input_dir) if f.endswith('.png')],
        key=lambda x: int(x.split('_')[1].split('.')[0])
    )

    def process_image(filename):
        image_path = os.path.join(input_dir, filename)
        t1 = extract_text_google_vision(image_path) # x=0, y=0
        t2 = extract_text_google_vision(image_path, x=1120) # x=1120, y=0
        t3 = extract_text_google_vision(image_path, x=1585) # x=1585, y=0
        return t1, t2, t3

    # Bilder parallel verarbeiten
    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(process_image, image_files))
        
    # Ergebnisse sammeln
    for t1, t2, t3 in results:
        ataChapter.append(t1)
        partNumber.append(t2)
        serialNumber.append(t3)

    # Daten in ein DataFrame umwandeln
    data = {
        'ataChapter': ataChapter,
        'partNumber': partNumber,
        'serialNumber': serialNumber
    }
    df = pd.DataFrame(data)

    # DataFrame in eine Excel-Datei speichern
    df.to_excel(output_file, index=False)
    print(f"Daten erfolgreich in Excel gespeichert unter {output_file}.")

# Aufruf der Funktion mit den zentral definierten Pfaden
extract_texts_to_excel_google_vision(output_dir_roi, os.path.join(output_dir_results, 'Ergebnis_mit_GoogleVisionAPI.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")


# Modul 9: Ergebnisse von Google Vision API formatieren

# Pfad zur Eingabe- und Ausgabedatei (gleiche Datei)
file_path = os.path.join(output_dir_results, 'Ergebnis_mit_GoogleVisionAPI.xlsx')

# Arbeitsmappe laden
wb = load_workbook(file_path)
sheet = wb.active

# Diese Funktion nimmt eine Zelle und überprüft, ob der Inhalt eine Liste in Textform ist.
# Falls die Zelle eine Liste in Form von Text enthält (z.B. "[ErsterWert, ZweiterWert]"),
# extrahiert die Funktion nur den ersten Wert und entfernt überflüssige Zeichen wie Klammern und Anführungszeichen.
def extract_first_value(cell):
    if isinstance(cell, str) and cell.startswith("[") and cell.endswith("]"):
        return cell.split(",")[0].strip().strip("'[")
    return cell

for row in sheet.iter_rows(min_row=2):  # Überspringe die Header-Zeile
    for cell in row:
        cell_value = cell.value
        new_value = extract_first_value(cell_value)
        cell.value = new_value

# Arbeitsmappe speichern
wb.save(file_path)
print(f"Formatiertes Excel gespeichert unter {file_path}.")


# Modul 10: Den innerhalb von Excel stattfindenden Qualitätstest vorbereiten, indem Ergebnisdateien mit Referenzdatei verknüpft werden

file_pairs = [
    (
        os.path.join(base_dir, 'Referenzdaten_0876_OCCM Status_GTR as of 22MAR24.xlsx'),
        os.path.join(output_dir_results, '1.Ergebnis_mit_Pytesseract.xlsx')
    ),
    (
        os.path.join(base_dir, 'Referenzdaten_0876_OCCM Status_GTR as of 22MAR24.xlsx'),
        os.path.join(output_dir_results, '2.Ergebnis_mit_Pytesseract.xlsx')
    ),
    (
        os.path.join(base_dir, 'Referenzdaten_0876_OCCM Status_GTR as of 22MAR24.xlsx'),
        os.path.join(output_dir_results, 'Ergebnis_mit_GoogleVisionAPI.xlsx')
    )
]

# Jedes Paar verarbeiten, um zusammengeführte Dateien zu erstellen
for original_file, ocr_file in file_pairs:
    # Daten aus beiden Dateien laden
    original_df = pd.read_excel(original_file)
    ocr_df = pd.read_excel(ocr_file)

    # Sicherstellen, dass beide DataFrames genau 3 Spalten haben
    if len(original_df.columns) != 3 or len(ocr_df.columns) != 3:
        print(f"Fehler: Eine der Dateien ({original_file}, {ocr_file}) hat nicht genau 3 Spalten.")
        continue

    # Spalten umbenennen für Klarheit bei der Zusammenführung
    merged_data = {
        'ataChapter_reference': original_df.iloc[:, 0],
        'partNumber_reference': original_df.iloc[:, 1],
        'serialNumber_reference': original_df.iloc[:, 2],
        'ataChapter_OCR': ocr_df.iloc[:, 0],
        'partNumber_OCR': ocr_df.iloc[:, 1],
        'serialNumber_OCR': ocr_df.iloc[:, 2],
    }

    # Zusammengeführtes DataFrame erstellen
    merged_df = pd.DataFrame(merged_data)

    # Ausgabedateinamen erstellen und speichern
    output_filename = os.path.join(output_dir_results, f'Qualitätstestdatei_{os.path.basename(ocr_file)}')
    merged_df.to_excel(output_filename, index=False)

    # Bestätigung, dass die Datei erstellt wurde
    print(f"Zusammengeführte Datei gespeichert: {output_filename}")


# Modul 11: Qualitätstest der extrahierten OCR-Daten durchführen

# Liste der zu verarbeitenden Dateien
file_paths = [
    os.path.join(output_dir_results, 'Qualitätstestdatei_1.Ergebnis_mit_Pytesseract.xlsx'),
    os.path.join(output_dir_results, 'Qualitätstestdatei_2.Ergebnis_mit_Pytesseract.xlsx'),
    os.path.join(output_dir_results, 'Qualitätstestdatei_Ergebnis_mit_GoogleVisionAPI.xlsx')
]

# Über jede Datei in der Liste iterieren
for file_path in file_paths:
    # Daten aus der Excel-Datei laden
    df = pd.read_excel(file_path)

    # Relevante Spalten in Strings umwandeln, um Vergleich mit Referenzdaten zu ermöglichen
    for col in ['ataChapter_reference', 'partNumber_reference', 'serialNumber_reference',
                'ataChapter_OCR', 'partNumber_OCR', 'serialNumber_OCR']:
        df[col] = df[col].astype(str)

    # Zähler für Gesamtzellen und Abweichungen initialisieren
    total_cells = 0
    mismatch_count = 0

    # Spaltenpaarweise vergleichen und Abweichungen zählen
    for col in ['ataChapter_reference', 'partNumber_reference', 'serialNumber_reference']:
        ocr_col = col + '_OCR'
        mismatches = df[col].str.strip() != df[ocr_col].str.strip()
        mismatch_count += mismatches.sum()
        total_cells += len(df[col])

    # Qualität berechnen
    accuracy = (total_cells - mismatch_count) / total_cells

    # WER (Word Error Rate) berechnen und zur Liste hinzufügen
    wer = (1 - accuracy) * 100
    wers.append(wer)
    
    
    # Arbeitsmappe laden, um Zellen farblich zu markieren
    wb = load_workbook(file_path)
    sheet = wb.active

    # Abweichende Zellen rot einfärben
    for index, row in df.iterrows():
        for col_idx, col_name in enumerate(['ataChapter_reference', 'partNumber_reference', 'serialNumber_reference']):
            original_cell = sheet.cell(row=index + 2, column=col_idx + 1)  # +2 wegen Header und 1-basierter Indexierung
            ocr_cell = sheet.cell(row=index + 2, column=col_idx + 4)
            if str(original_cell.value).strip() != str(ocr_cell.value).strip():
                ocr_cell.font = Font(color="FF0000")

    # Arbeitsmappe speichern
    wb.save(file_path)

    # Genauigkeiten und WER für die aktuelle ausgeben
    print(f"Genauigkeit der OCR-Daten in {file_paths}: {accuracy*100:.2f}%")
    print(f"Word Error Rate (WER) der OCR-Daten in {file_paths}: {wer:.2f}%")