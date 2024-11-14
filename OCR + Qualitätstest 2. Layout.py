# Modul 1: Import von Bibliotheken und Definition der Pfade

import time  # Zeitmessung der Zellausführung
import cv2  # OpenCV-Bibliothek für Bildverarbeitung
import numpy as np  # Für mathematische Operationen mit Arrays
import random  # Zum Generieren von Zufallszahlen
import os  # Um mit dem Dateisystem zu arbeiten
from pdf2image import convert_from_path, pdfinfo_from_path  # Um PDF-Seiten in Bilder umzuwandeln
import gc  # Für die Speicherbereinigung
from PIL import Image
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

import pytesseract  # Open Source Text-OCR Software
from google.cloud import vision

# Zentral definierte Pfade und Einstellungen

# Basisverzeichnis für das Projekt
base_dir = 'C:/Users/Katharina/OCR - PDF_to_Excel/2. PDF Lufthansa/' # Hier den Ordnerpfad angeben indem die PDF liegt. Außer der PDF sollte keine der anderen PDF im gleichen Ordner liegen sodass die sich die später erstellten Output-Verzeichnisse nicht bei gleichem Namen überschreiben.

# Eingabedatei (PDF)
input_pdf_file = os.path.join(base_dir, 'Aircraft Equipment Liste Report_D-AIHE.pdf')

# Output-Verzeichnisse
output_dir_png = os.path.join(base_dir, '1_PDF_to_PNG')
output_dir_cropped = os.path.join(base_dir, '2_zugeschnittene PNG')
output_dir_results = os.path.join(base_dir, '3_Ergebnisse')

# Referenzdatei für den Genauigkeitstest (mit manuell überprüften 100% richtigen Werten) 
referenzdatei = os.path.join(base_dir, 'Referenzdaten_Aircraft Equipment Liste Report_D-AIHE.xlsx')

# Pfad zur Tesseract-OCR-Programmdatei (Installationstutorial: https://www.youtube.com/watch?v=O8maBz1yXe0 Downloadlink (für Windows): https://github.com/UB-Mannheim/tesseract/wiki)
tesseract_cmd = r'C:\Users\Katharina\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'

# Pfad zur lokalen Google Vision API JSON-Datei (Installationstutorial: https://www.youtube.com/watch?v=OFrwjBbk9SY bis Minute 2:10)
google_vision_credentials = 'C:/Users/Katharina/OCR - PDF_to_Excel/vision-api.json' 

# Pfad zum Poppler \Library\bin Ordner (Download/Installtion: https://www.geeksforgeeks.org/convert-pdf-to-image-using-python/)
poppler_path = r"C:\Users\Katharina\poppler-24.07.0\Library\bin"

# Sicherstellen, dass die Output-Verzeichnisse existieren
os.makedirs(output_dir_png, exist_ok=True)
os.makedirs(output_dir_cropped, exist_ok=True)
os.makedirs(output_dir_results, exist_ok=True)

# Modul 2: Jede PDF-Seite in eine PNG-Datei umwandeln und speichern (mit Hilfe von PDF2Image und Poppler)

start_time = time.time()

def extract_pdf_to_png(pdf_file, output_folder, dpi=300): 
    # Gesamtanzahl der Seiten ermitteln
    info = pdfinfo_from_path(
        pdf_file,
        poppler_path=poppler_path
    )
    max_pages = info["Pages"]

    # Seiten einzeln verarbeiten, um Speicherüberlastung zu vermeiden
    for page_number in range(1, max_pages + 1):
        images = convert_from_path(
            pdf_file,
            dpi=dpi,
            first_page=page_number,
            last_page=page_number,
            poppler_path=poppler_path
        )
        for image in images:
            output_path = os.path.join(output_folder, f'page_{page_number}.png')
            image.save(output_path, 'PNG')
            print(f"Seite {page_number} gespeichert unter {output_path}")

            # Speicher für jede Seite explizit freigeben
            del image
            gc.collect()  # Speicherbereinigung durchführen

        del images
        gc.collect()

    print(f"Alle Seiten wurden erfolgreich in {output_folder} gespeichert.")

# Aufruf der Funktion mit den zentral definierten Pfaden
extract_pdf_to_png(input_pdf_file, output_dir_png)

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 3: Die Seitenbilder auf die relevanten Tabellenstellen zuschneiden (verkleinern) und erneut abspeichern

start_time = time.time()

# Alle PNG-Dateien im Eingabeordner sammeln und sortieren
image_files = sorted([f for f in os.listdir(output_dir_png) if f.endswith('.png')])

# Erste Seite verarbeiten
first_page = image_files[0]
image_path = os.path.join(output_dir_png, first_page)
image = cv2.imread(image_path)

# Bereich für die erste Seite zuschneiden
left_half = image[976:2430, 60:1380]

# Geschnittenes Bild speichern
output_path = os.path.join(output_dir_cropped, first_page)
cv2.imwrite(output_path, left_half)

# Restliche Seiten verarbeiten
for filename in image_files[1:]:
    image_path = os.path.join(output_dir_png, filename)
    image = cv2.imread(image_path)

    # Bereich für die folgenden Seiten zuschneiden
    left_half = image[325:2430, 60:1380]

    # Geschnittenes Bild speichern
    output_path = os.path.join(output_dir_cropped, filename)
    cv2.imwrite(output_path, left_half)
            
print("Verarbeitung abgeschlossen. Bilder gespeichert in", output_dir_cropped)

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 4: In den klein zugeschnittenen Seitenbildern alle horizontalen Tabellenlinien erkennen, welche sich direkt ober- und unterhalb von den relevanten Teiledaten befinden. 

def detect_horizontal_lines(image_path, line_width_percentage=0.9):
    image = cv2.imread(image_path)
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # Das Bild in Graustufen umwandeln
    
    # Kanten erkennen
    edges = cv2.Canny(image, 50, 150, apertureSize=3)
    
    # Hough-Transformation zur Linienerkennung
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, 100, minLineLength=100, maxLineGap=10)
    
    if lines is None:
        return []
    
    horizontal_lines = []
    img_height, img_width = image.shape[:2]
    line_width = int(img_width * line_width_percentage)

    # Nur horizontale Linien mit ausreichender Länge behalten
    for line in lines:
        x1, y1, x2, y2 = line[0]
        if abs(x2 - x1) > line_width:
            horizontal_lines.append((y1, y2))

    return horizontal_lines

def draw_lines_on_image(image_path, lines):
    image = cv2.imread(image_path)
    # Linien auf dem Bild zeichnen
    for y1, y2 in lines:
        cv2.line(image, (0, y1), (image.shape[1], y1), (0, 255, 0), 2)
    return image

# Modul 5: Funktion, um Teiledaten zwischen den erkannten horizontalen Linien an hartkodierter x-Koordinate mit Pytesseract zu extrahieren

# Pfad zu Tesseract festlegen
pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

def extract_text_from_lines(image_path, lines, vertical_distance_threshold):
    image = cv2.imread(image_path)
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # Das Bild in Graustufen umwandeln
    texts = []
   
    # Linien von oben nach unten sortieren
    sorted_lines = sorted(lines, key=lambda x: x[0])
    for i in range(len(sorted_lines) - 1):
        y1 = sorted_lines[i][0]
        y2 = sorted_lines[i + 1][0]
        
        # Prüfen, ob der Abstand zwischen den Linien passt
        if (y2 - y1) > vertical_distance_threshold and (y2 - y1) < 90:
            # Region of Interest (ROI) für verschiedene Spalten
            roi = image[y1:y2, 10:125]
            text1 = pytesseract.image_to_string(roi, config='--psm 6').strip()
            
            roi = image[y1:y2, 565:945]
            text2 = pytesseract.image_to_string(roi, config='--psm 6').strip()
            text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
            
            roi = image[y1:y2, 955:1310]
            text3 = pytesseract.image_to_string(roi, config='--psm 6').strip()
            
            if text1 or text2 or text3:
                texts.append((text1, text2, text3))
    
    # Spezielle Behandlung für bestimmte Seiten aufgrund unregelmäßiger Tabellenform
    filename = os.path.basename(image_path)
    page_no = int(filename.split('_')[1].split('.')[0])
    if page_no in (19, 50, 138):
        y1 = 1 
        y2 = 55
        roi = image[y1:y2, 10:125]
        text1 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        roi = image[y1:y2, 565:945]
        text2 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 955:1310]
        text3 = pytesseract.image_to_string(roi, config='--psm 6').strip()
       
        texts.insert(0, (text1, text2, text3))
    elif page_no == 51:
        y1 = 575 
        y2 = 625
        roi = image[y1:y2, 10:125]
        text1 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        roi = image[y1:y2, 565:945]
        text2 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 955:1310]
        text3 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        texts.insert(6, (text1, text2, text3))
    elif page_no == 137:
        y1 = 760 
        y2 = 810
        roi = image[y1:y2, 10:125]
        text1 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        roi = image[y1:y2, 570:945]
        text2 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 955:1310]
        text3 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        texts.insert(8, (text1, text2, text3))

        y1 = 1120 
        y2 = 1170
        roi = image[y1:y2, 10:125]
        text1 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        roi = image[y1:y2, 565:945]
        text2 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 955:1310]
        text3 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        texts.insert(11, (text1, text2, text3))
    elif page_no == 199:
        y1 = 1870
        y2 = 1920
        roi = image[y1:y2, 10:125]
        text1 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        roi = image[y1:y2, 570:945]
        text2 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 960:1310]
        text3 = pytesseract.image_to_string(roi, config='--psm 6').strip()
        
        texts.append((text1, text2, text3))

    print(f"Verarbeitung abgeschlossen für: {image_path}")
    return texts

# Modul 6: Daten/Text-Extraktion ausführen und in 1. Exceldatei speichern - 1. Ergebnis ohne Eliminierug von häufigsten Fehlern

start_time = time.time()

def extract_texts_to_excel(input_dir, output_file):
    ataChapter = []
    partNumber = []
    serialNumber = []
    image_files = sorted([f for f in os.listdir(input_dir) if f.endswith('.png')], 
                         key=lambda x: int(x.split('_')[1].split('.')[0]))
    print(image_files)
    
    # Funktion zur Verarbeitung eines Bildes
    def process_image(filename):
        image_path = os.path.join(input_dir, filename)
        
        horizontal_lines = detect_horizontal_lines(image_path) 
        texts = extract_text_from_lines(image_path, horizontal_lines, vertical_distance_threshold=50)

        # Ergebnisse sammeln
        if texts:
            ata, part_no, serial_no = zip(*texts)
            return list(ata), list(part_no), list(serial_no)
        else:
            return [], [], []
    
    # Bilder parallel verarbeiten
    with ThreadPoolExecutor() as executor:
        results = list(executor.map(process_image, image_files))

    # Ergebnisse aus allen Bildern zusammenführen
    for result in results:
        ata, part_no, serial_no = result
        ataChapter.extend(ata)
        partNumber.extend(part_no)
        serialNumber.extend(serial_no)
    
    # Daten in ein Dictionary packen
    data = {
        'ATA': ataChapter,
        'PORT NO.': partNumber,
        'SERIAL NO.': serialNumber
    }
    
    # In DataFrame umwandeln
    df = pd.DataFrame(data)

    # Leere Einträge mit NaN ersetzen und auffüllen
    df['ATA'] = df['ATA'].replace('', np.nan)
    df['ATA'] = df['ATA'].ffill()

    # Ungewünschte Zeilen entfernen
    df = df[df['PORT NO.'] != 'REQUIREMENT']
    
    # DataFrame in Excel speichern
    df.to_excel(output_file, index=False)
    print("Daten erfolgreich in Excel gespeichert.")

# Funktion aufrufen
extract_texts_to_excel(output_dir_cropped, os.path.join(output_dir_results, '1.Ergebnis_mit_Pytesseract.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 7: Daten/Text-Extraktion ausführen und in 2. Exceldatei speichern - 2. Ergebnis MIT Eliminierug von häufigsten Fehlern (hartkodiert)

start_time = time.time()

def extract_texts_to_excel(input_dir, output_file):
    ataChapter = []
    partNumber = []
    serialNumber = []
    image_files = sorted([f for f in os.listdir(input_dir) if f.endswith('.png')], 
                         key=lambda x: int(x.split('_')[1].split('.')[0]))
    print(image_files)
    
    # Funktion zur Verarbeitung eines Bildes
    def process_image(filename):
        image_path = os.path.join(input_dir, filename)
        
        horizontal_lines = detect_horizontal_lines(image_path) 
        texts = extract_text_from_lines(image_path, horizontal_lines, vertical_distance_threshold=50)

        # Ergebnisse sammeln
        if texts:
            ata, part_no, serial_no = zip(*texts)
            return list(ata), list(part_no), list(serial_no)
        else:
            return [], [], []
    
    # Bilder parallel verarbeiten
    with ThreadPoolExecutor() as executor:
        results = list(executor.map(process_image, image_files))

    # Ergebnisse aus allen Bildern zusammenführen
    for result in results:
        ata, part_no, serial_no = result
        ataChapter.extend(ata)
        partNumber.extend(part_no)
        serialNumber.extend(serial_no)
    
    # Daten in ein Dictionary packen
    data = {
        'ATA': ataChapter,
        'PORT NO.': partNumber,
        'SERIAL NO.': serialNumber
    }
    
    # In DataFrame umwandeln
    df = pd.DataFrame(data)

    # Leerzeichen entfernen und "$" durch "S" ersetzen
    df = df.applymap(lambda x: x.replace("$", "S") if isinstance(x, str) else x)

    # Leere Einträge mit NaN ersetzen und auffüllen
    df['ATA'] = df['ATA'].replace('', np.nan)
    df['ATA'] = df['ATA'].ffill()

    # Ungewünschte Zeilen entfernen
    df = df[df['PORT NO.'] != 'REQUIREMENT']
    
    # DataFrame in Excel speichern
    df.to_excel(output_file, index=False)
    print("Daten erfolgreich in Excel gespeichert.")

# Funktion aufrufen
extract_texts_to_excel(output_dir_cropped, os.path.join(output_dir_results, '2.Ergebnis_mit_Pytesseract.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 8: Google Vision API und Funktion einrichten, um Teiledaten zwischen den erkannten horizontalen Linien an hartkodierter x-Koordinate zu extrahieren

# Anmeldeinformationen für die Google Vision API festlegen
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = google_vision_credentials
client = vision.ImageAnnotatorClient()

def get_text_google_ocr(roi):
    success, encoded_image = cv2.imencode('.png', roi)
    content = encoded_image.tobytes()
    image = vision.Image(content=content)
    response = client.text_detection(image=image)
    texts = response.text_annotations
    full_text = []
    for text in texts:
        full_text.append(text.description)
    if len(full_text):
        full_text = full_text[0]
    else:
        full_text = ""
    return full_text

def extract_text_from_lines(image_path, lines, vertical_distance_threshold):
    time.sleep(2)
    image = cv2.imread(image_path)
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    texts = []

    # Linien sortieren
    sorted_lines = sorted(lines, key=lambda x: x[0])
    for i in range(len(sorted_lines) - 1):
        y1 = sorted_lines[i][0]
        y2 = sorted_lines[i + 1][0]

        # Prüfen, ob der Abstand passt
        if (y2 - y1) > vertical_distance_threshold and (y2 - y1) < 90:
            roi = image[y1:y2, 10:125]
            text1 = get_text_google_ocr(roi)
            
            roi = image[y1:y2, 565:945]
            text2 = get_text_google_ocr(roi)
            text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
            
            roi = image[y1:y2, 955:1310]
            text3 = get_text_google_ocr(roi)
            
            if text1 or text2 or text3:
                texts.append((text1, text2, text3))

    # Bestimmte Seiten speziell behandeln aufgrund unregelmäßiger Tabellenform
    filename = os.path.basename(image_path)
    page_no = int(filename.split('_')[1].split('.')[0])
    if page_no in (19, 50, 138):
        y1 = 1
        y2 = 55
        roi = image[y1:y2, 10:125]
        text1 = get_text_google_ocr(roi)
        
        roi = image[y1:y2, 565:945]
        text2 = get_text_google_ocr(roi)
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 955:1310]
        text3 = get_text_google_ocr(roi)

        texts.insert(0, (text1, text2, text3))
    elif page_no == 51:
        y1 = 575
        y2 = 625
        roi = image[y1:y2, 10:125]
        text1 = get_text_google_ocr(roi)
        
        roi = image[y1:y2, 565:945]
        text2 = get_text_google_ocr(roi)
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 955:1310]
        text3 = get_text_google_ocr(roi)

        texts.insert(6, (text1, text2, text3))
    elif page_no == 137:
        y1 = 760
        y2 = 810
        roi = image[y1:y2, 10:125]
        text1 = get_text_google_ocr(roi)
        
        roi = image[y1:y2, 570:945]
        text2 = get_text_google_ocr(roi)
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 955:1310]
        text3 = get_text_google_ocr(roi)

        texts.insert(8, (text1, text2, text3))

        y1 = 1120
        y2 = 1170
        roi = image[y1:y2, 10:125]
        text1 = get_text_google_ocr(roi)
        
        roi = image[y1:y2, 565:945]
        text2 = get_text_google_ocr(roi)
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 955:1310]
        text3 = get_text_google_ocr(roi)

        texts.insert(11, (text1, text2, text3))
    elif page_no == 199:
        y1 = 1870
        y2 = 1920
        roi = image[y1:y2, 10:125]
        text1 = get_text_google_ocr(roi)
        
        roi = image[y1:y2, 570:945]
        text2 = get_text_google_ocr(roi)
        text2 = re.sub(r'^UNKNOWN_STRU.*', 'UNKNOWN_STRUCTUR', text2)
        
        roi = image[y1:y2, 960:1310]
        text3 = get_text_google_ocr(roi)

        texts.append((text1, text2, text3))

    print(f"Verarbeitung abgeschlossen für: {image_path}")
    return texts

# Modul 9: Daten/Text-Extraktion mit Google Vision API ausführen und in 3. Exceldatei speichern

start_time = time.time()

def extract_texts_to_excel(input_dir, output_file):
    ataChapter = []
    partNumber = []
    serialNumber = []

    # Bilddateien sammeln und sortieren
    image_files = sorted(
        [f for f in os.listdir(input_dir) if f.endswith('.png')],
        key=lambda x: int(x.split('_')[1].split('.')[0])
    )
    print(image_files)

    def process_image(filename):
        image_path = os.path.join(input_dir, filename)

        # Horizontale Linien erkennen
        horizontal_lines = detect_horizontal_lines(image_path)

        # Text extrahieren
        texts = extract_text_from_lines(image_path, horizontal_lines, vertical_distance_threshold=50)

        # Ergebnisse sammeln
        if texts:
            ata, part_no, serial_no = zip(*texts)
            return list(ata), list(part_no), list(serial_no)
        else:
            return [], [], []

    # Bilder parallel verarbeiten (max. 2 gleichzeitig)
    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(process_image, image_files))

    # Ergebnisse zusammenführen
    for result in results:
        ata, part_no, serial_no = result
        ataChapter.extend(ata)
        partNumber.extend(part_no)
        serialNumber.extend(serial_no)

    # Daten in DataFrame umwandeln
    data = {
        'ATA': ataChapter,
        'PORT NO.': partNumber,
        'SERIAL NO.': serialNumber
    }
    df = pd.DataFrame(data)

    # Leere Einträge mit NaN ersetzen und auffüllen
    df['ATA'] = df['ATA'].replace('', np.nan)
    df['ATA'] = df['ATA'].ffill()

    # Ungewünschte Zeilen entfernen
    df = df[df['PORT NO.'] != 'REQUIREMENT']

    # DataFrame in Excel speichern
    df.to_excel(output_file, index=False)
    print("Daten erfolgreich in Excel gespeichert.")

# Funktion aufrufen
extract_texts_to_excel(output_dir_cropped, os.path.join(output_dir_results, 'Ergebnis_mit_GoogleVisionAPI.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Zelle 10: Vergleichsdateien (Referenzdaten+OCR Ergebnisse) für Qualitätstest erstellen

# Liste von Dateipaaren definieren, die zusammengeführt werden sollen
file_pairs = [
    (
        referenzdatei,
        os.path.join(output_dir_results, '2.Ergebnis_mit_Pytesseract.xlsx')
    ),
    (
        referenzdatei,
        os.path.join(output_dir_results, '1.Ergebnis_mit_Pytesseract.xlsx')
    ),
    (
        referenzdatei,
        os.path.join(output_dir_results, 'Ergebnis_mit_GoogleVisionAPI.xlsx')
    )
]

# Ausgabeverzeichnis ist bereits definiert: output_dir_results

# Jedes Paar verarbeiten, um zusammengeführte Dateien zu erstellen
for original_file, ocr_file in file_pairs:
    # Daten aus beiden Dateien laden
    original_df = pd.read_excel(original_file)
    ocr_df = pd.read_excel(ocr_file)

    # Sicherstellen, dass beide DataFrames genau 3 Spalten haben
    if len(original_df.columns) != 3 or len(ocr_df.columns) != 3:
        print(f"Fehler: Eine der Dateien ({original_file}, {ocr_file}) hat nicht genau 3 Spalten.")
        continue

    # Spalten umbenennen für Klarheit bei der Zusammenführung
    merged_data = {
        'ataChapter': original_df.iloc[:, 0],
        'partNumber': original_df.iloc[:, 1],
        'serialNumber': original_df.iloc[:, 2],
        'ataChapter_OCR': ocr_df.iloc[:, 0],
        'partNumber_OCR': ocr_df.iloc[:, 1],
        'serialNumber_OCR': ocr_df.iloc[:, 2],
    }

    # Zusammengeführtes DataFrame erstellen
    merged_df = pd.DataFrame(merged_data)

    # Ausgabedateinamen erstellen und speichern
    output_filename = os.path.join(output_dir_results, f'Genauigkeitstestdatei_{os.path.basename(ocr_file)}')
    merged_df.to_excel(output_filename, index=False)

    # Bestätigung, dass die Datei erstellt wurde
    print(f"Zusammengeführte Datei gespeichert: {output_filename}")

# Modul 11: Qualitätstest der extrahierten OCR-Daten durchführen

# Liste der zu verarbeitenden Dateien
file_paths = [
    os.path.join(output_dir_results, 'Qualitätstestdatei_2.Ergebnis_mit_Pytesseract.xlsx'),
    os.path.join(output_dir_results, 'Qualitätstestdatei_1.Ergebnis_mit_Pytesseract.xlsx'),
    os.path.join(output_dir_results, 'Qualitätstestdatei_Ergebnis_mit_GoogleVisionAPI.xlsx')
]

# Über jede Datei in der Liste iterieren
for file_path in file_paths:
    # Daten aus der Excel-Datei laden
    df = pd.read_excel(file_path)
    
    # Relevante Spalten in Strings umwandeln, um Vergleich mit Referenzdaten zu ermöglichen
    for col in ['ataChapter', 'partNumber', 'serialNumber',
                'ataChapter_OCR', 'partNumber_OCR', 'serialNumber_OCR']:
        df[col] = df[col].astype(str)
    
    # Zähler für Gesamtzellen und Abweichungen initialisieren
    total_cells = 0
    mismatch_count = 0
    
    # Spaltenpaarweise vergleichen und Abweichungen zählen
    for col in ['ataChapter', 'partNumber', 'serialNumber']:
        ocr_col = col + '_OCR'
        mismatches = df[col].str.strip() != df[ocr_col].str.strip()
        mismatch_count += mismatches.sum()
        total_cells += len(df[col])
    
    # Qualität berechnen
    accuracy = (total_cells - mismatch_count) / total_cells

    # WER (Word Error Rate) berechnen und zur Liste hinzufügen
    wer = (1 - accuracy) * 100
    wers.append(wer)
    
    # Arbeitsmappe laden, um Zellen farblich zu markieren
    wb = load_workbook(file_path)
    sheet = wb.active
    
    # Abweichende Zellen rot einfärben
    for index, row in df.iterrows():
        for col_idx, col_name in enumerate(['ataChapter', 'partNumber', 'serialNumber']):
            original_cell = sheet.cell(row=index + 2, column=col_idx + 1)  # +2 wegen Header und 1-basierter Indexierung
            ocr_cell = sheet.cell(row=index + 2, column=col_idx + 4)
            if str(original_cell.value).strip() != str(ocr_cell.value).strip():
                ocr_cell.font = Font(color="FF0000")
    
    # Arbeitsmappe speichern
    wb.save(file_path)
    
# Genauigkeiten und WER für die aktuelle ausgeben
    print(f"Genauigkeit der OCR-Daten in {file_paths}: {accuracy*100:.2f}%")
    print(f"Word Error Rate (WER) der OCR-Daten in {file_paths}: {wer:.2f}%")

