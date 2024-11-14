# Modul 1: Import von Bibliotheken und Definition der Pfade

import time  # Zeitmessung der Zellausführung
import cv2  # OpenCV-Bibliothek für Bildverarbeitung
import numpy as np  # Für mathematische Operationen mit Arrays
import random  # Zum Generieren von Zufallszahlen
import os  # Um mit dem Dateisystem zu arbeiten
from pdf2image import convert_from_path, pdfinfo_from_path  # Um PDF-Seiten in Bilder umzuwandeln
import gc  # Garbage Collector zum Freigeben von Speicher
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import Font

import pytesseract  # Open Source Text-OCR Software
from google.cloud import vision

# Zentral definierte Pfade und Einstellungen

# Basisverzeichnis für das Projekt (In diesem Verzeichnis sollte die PDF gespeichert sein und gegebenfalls die 100% korrekte Referenzdatei falls man einen Qualitätstest durchführen möchte)
base_dir = 'C:/Users/Katharina/OCR - PDF_to_Excel/3. PDF Skytraders/'

# Eingabedatei (PDF)
input_pdf_file = os.path.join(base_dir, 'N-01999_2016-09-02_Hard Time Component Status_MSN1999_HT Component Status.pdf')

# Output-Verzeichnisse
output_dir_png = os.path.join(base_dir, '1_PDF_to_PNG')
output_dir_cropped = os.path.join(base_dir, '2_zugeschnittene PNG')
output_dir_results = os.path.join(base_dir, '3_Ergebnisse')

# Referenzdatei
referenzdatei = os.path.join(base_dir, 'Referenzdaten_N-01999_2016-09-02_Hard Time Component Status_MSN1999_HT Component Status.xlsx')

# Pfad zur Tesseract-OCR-Programmdatei (Installationstutorial: https://www.youtube.com/watch?v=O8maBz1yXe0 Downloadlink (für Windows): https://github.com/UB-Mannheim/tesseract/wiki)
tesseract_cmd = r'C:\Users\Katharina\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'
pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

# Pfad zur lokalen Google Vision API JSON-Datei (Installationstutorial: https://www.youtube.com/watch?v=OFrwjBbk9SY bis Minute 2:10)
google_vision_credentials = 'C:/Users/Katharina/OCR - PDF_to_Excel/bachelorarbeit-441110-7a84ba3a3b1c.json'
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = google_vision_credentials

# Sicherstellen, dass die Output-Verzeichnisse existieren
os.makedirs(output_dir_png, exist_ok=True)
os.makedirs(output_dir_cropped, exist_ok=True)
os.makedirs(output_dir_results, exist_ok=True)

# Modul 2: Jede PDF-Seite in eine PNG-Datei umwandeln und speichern (mit Hilfe von PDF2Image und Poppler)

start_time = time.time()

def extract_pdf_to_png(pdf_file, output_folder, dpi=300):
    # Erstelle den Ausgabeordner, falls er nicht existiert
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Jede Seite des PDFs in ein Bild umwandeln und speichern
    for i, page in enumerate(convert_from_path(pdf_file, dpi=dpi)):
        output_path = os.path.join(output_folder, f'page_{i + 1}.png')
        page.save(output_path, 'PNG')
        print(f"Seite {i + 1} gespeichert unter {output_path}")

        # Speicher freigeben
        del page
        gc.collect()

    print(f"Alle Seiten wurden erfolgreich in {output_folder} gespeichert.")

# Funktion aufrufen, um Bilder aus dem PDF zu extrahieren
extract_pdf_to_png(input_pdf_file, output_dir_png)

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 3: Die Seitenbilder zuschneiden (verkleinern) und erneut abspeichern

start_time = time.time()

# Liste der Bilddateien im Eingabeverzeichnis sortiert laden
image_files = sorted(
    [f for f in os.listdir(output_dir_png) if f.endswith('.png')],
    key=lambda x: int(x.split('_')[1].split('.')[0]))
print(image_files)

# Erste Seite überspringen und die restlichen verarbeiten
for filename in image_files[1:]:
    if filename.endswith('.png'):
        image_path = os.path.join(output_dir_png, filename)
        image = cv2.imread(image_path)

        # Bild ab einer bestimmten Höhe zuschneiden
        left_half = image[320:, :]

        # Zugeschnittenes Bild im Ausgabeordner speichern
        output_path = os.path.join(output_dir_cropped, filename)
        cv2.imwrite(output_path, left_half)

print("Verarbeitung abgeschlossen. Bilder gespeichert in", output_dir_cropped)

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 4: In den klein zugeschnittenen Seitenbildern alle horizontalen Linien erkennen, welche sich alle Teile voneinander abgrenzen.

def detect_horizontal_lines(image_path, line_width_percentage=0.9):
    image = cv2.imread(image_path)
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Kanten im Bild erkennen
    edges = cv2.Canny(image, 50, 150, apertureSize=3)

    # Hough-Transformation zur Linienerkennung anwenden
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, 100, minLineLength=100, maxLineGap=10)

    # Falls keine Linien erkannt wurden, leere Liste zurückgeben
    if lines is None:
        return []

    horizontal_lines = []
    img_height, img_width = image.shape[:2]
    line_width = int(img_width * line_width_percentage)

    # Über die erkannten Linien iterieren
    for line in lines:
        x1, y1, x2, y2 = line[0]
        # Prüfen, ob die Linie horizontal ist und länger als die minimale Breite
        if abs(y2 - y1) < 10 and abs(x2 - x1) > line_width:
            horizontal_lines.append((y1, y2))

    return horizontal_lines

def draw_lines_on_image(image_path, lines):
    image = cv2.imread(image_path)
    # Jede horizontale Linie auf das Bild zeichnen
    for y1, y2 in lines:
        cv2.line(image, (0, y1), (image.shape[1], y1), (0, 255, 0), 2)

    return image

# Modul 5: Funktion, um Teiledaten zwischen den erkannten horizontalen Linien an hartkodierter x und y Koordinatenstelle mit Pytesseract zu extrahieren

def extract_text_between_lines(image_path, horizontal_lines, distance_threshold=120):
    extracted_texts = []

    image = cv2.imread(image_path)
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # Sortieren der horizontalen Linien nach ihrer y-Koordinate
    horizontal_lines = sorted(horizontal_lines, key=lambda x: x[0])

    for i in range(len(horizontal_lines) - 1):
        y1 = horizontal_lines[i][0]
        y2 = horizontal_lines[i+1][0]
        
        # Prüfen, ob der Abstand zwischen den Linien größer als der Schwellenwert ist
        if y2 - y1 > distance_threshold:
            # Bereich zwischen den beiden Linien ausschneiden
            text_region = image[y1:y2-125, 110:370]
            # Text mit PyTesseract extrahieren
            ata = pytesseract.image_to_string(text_region).strip()

            text_region = image[y1+60:y2-40, 560:850]
            # Part Number und Serial Number extrahieren
            extracted_text = pytesseract.image_to_string(text_region).strip().split('\n')
            if len(extracted_text) >= 2:
                part_no, serial_no = extracted_text[:2]
            else:
                part_no = extracted_text[0] if extracted_text else ''
                serial_no = ''
            extracted_texts.append((ata, part_no, serial_no))
    print(f"Verarbeitung abgeschlossen für: {image_path}")
    return extracted_texts

# Modul 6: Daten/Text-Extraktion ausführen und in 1. Exceldatei speichern - 1. Ergebnis ohne Eliminierug von häufigsten Fehlern

start_time = time.time()

def extract_texts_to_excel(input_dir, output_file):
    ataChapter = []
    partNumber = []
    serialNumber = []
    image_files = sorted([f for f in os.listdir(input_dir) if f.endswith('.png')], 
                         key=lambda x: int(x.split('_')[1].split('.')[0]))

    print(image_files)
    for filename in image_files:
        image_path = os.path.join(input_dir, filename)
        
        horizontal_lines = detect_horizontal_lines(image_path) 
        texts = extract_text_between_lines(image_path, horizontal_lines)

        # Überprüfen, ob Texte extrahiert wurden
        if texts:
            # Tupel in drei separate Listen entpacken
            ata, part_no, serial_no = zip(*texts)

            ataChapter += (list(ata))
            partNumber += (list(part_no))
            serialNumber += (list(serial_no))   
        else:
            print(f"Keine Texte in {filename} extrahiert.")
        
    # Daten in ein Dictionary packen
    data = {
        'ATA': ataChapter,
        'PORT NO.': partNumber,
        'SERIAL NO.': serialNumber
    }

    # In DataFrame umwandeln
    df = pd.DataFrame(data)

    # DataFrame in Excel speichern
    df.to_excel(output_file, index=False)
    print("Daten erfolgreich in Excel gespeichert.")

# Funktion anwenden, um alle Bilder zu verarbeiten und Ergebnisse zu speichern
extract_texts_to_excel(output_dir_cropped, os.path.join(output_dir_results, '1.Ergebnis_mit_Pytesseract.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 7: Daten/Text-Extraktion ausführen und in 2. Exceldatei speichern - 2. Ergebnis MIT Eliminierug von häufigsten Fehlern (hartkodiert)

start_time = time.time()

def extract_texts_to_excel(input_dir, output_file):
    ataChapter = []
    partNumber = []
    serialNumber = []
    image_files = sorted([f for f in os.listdir(input_dir) if f.endswith('.png')], 
                         key=lambda x: int(x.split('_')[1].split('.')[0]))

    print(image_files)
    for filename in image_files:
        image_path = os.path.join(input_dir, filename)
        
        horizontal_lines = detect_horizontal_lines(image_path) 
        texts = extract_text_between_lines(image_path, horizontal_lines)

        # Überprüfen, ob Texte extrahiert wurden
        if texts:
            # Tupel in drei separate Listen entpacken
            ata, part_no, serial_no = zip(*texts)

            ataChapter += (list(ata))
            partNumber += (list(part_no))
            serialNumber += (list(serial_no))   
        else:
            print(f"Keine Texte in {filename} extrahiert.")
        
    # Daten in ein Dictionary packen
    data = {
        'ATA': ataChapter,
        'PORT NO.': partNumber,
        'SERIAL NO.': serialNumber
    }

    # In DataFrame umwandeln
    df = pd.DataFrame(data)
    df = df.applymap(lambda x: x.replace('$', 'S') if isinstance(x, str) else x)

    # DataFrame in Excel speichern
    df.to_excel(output_file, index=False)
    print("Daten erfolgreich in Excel gespeichert.")

# Funktion anwenden, um alle Bilder zu verarbeiten und Ergebnisse zu speichern
extract_texts_to_excel(output_dir_cropped, os.path.join(output_dir_results, '2.Ergebnis_mit_Pytesseract.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 8: Google Vision API und Funktion einrichten, um Teiledaten zwischen den erkannten horizontalen Linien an hartkodierter x und y-Koordinate zu extrahieren

client = vision.ImageAnnotatorClient()

def get_google_ocr(roi):
    success, encoded_image = cv2.imencode('.png', roi)
    content = encoded_image.tobytes()
    image = vision.Image(content=content)
    response = client.text_detection(image=image)
    texts = response.text_annotations
    full_text = []
    for text in texts:
        full_text.append(text.description)

    if len(full_text):
        full_text = full_text[0].split("\n")
    return full_text

def extract_text_between_lines(image_path, horizontal_lines, distance_threshold=120):
    time.sleep(1)
    extracted_texts = []

    image = cv2.imread(image_path)
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Sortieren der horizontalen Linien von oben nach unten
    horizontal_lines = sorted(horizontal_lines, key=lambda x: x[0])

    # Über Paare benachbarter Linien iterieren
    for i in range(len(horizontal_lines) - 1):
        y1 = horizontal_lines[i][0]
        y2 = horizontal_lines[i+1][0]

        # Prüfen, ob der vertikale Abstand zwischen den Linien größer als der Schwellenwert ist
        if y2 - y1 > distance_threshold:
            # Bereich zwischen den Linien für OCR extrahieren
            text_region_ata = image[y1:y2-125, 110:370]
            ata = get_google_ocr(text_region_ata)

            text_region_part_serial = image[y1+60:y2-40, 560:850]
            # Part Number und Serial Number extrahieren
            part_serial_text = get_google_ocr(text_region_part_serial)
            if len(part_serial_text) >= 2:
                part_no, serial_no = part_serial_text[:2]
            else:
                part_no = part_serial_text[0] if part_serial_text else ''
                serial_no = ''
            extracted_texts.append((ata, part_no, serial_no))

    print(f"Verarbeitung abgeschlossen für: {image_path}")
    return extracted_texts

# Modul 9: Daten/Text-Extraktion mit Google Vision API ausführen und in 3. Exceldatei speichern

start_time = time.time()

def extract_texts_to_excel_google(input_dir, output_file):
    ataChapter = []
    partNumber = []
    serialNumber = []

    # Sortierte Liste der Bilddateien im Eingabeverzeichnis
    image_files = sorted(
        [f for f in os.listdir(input_dir) if f.endswith('.png')],
        key=lambda x: int(x.split('_')[1].split('.')[0])
    )

    print(image_files)
    for filename in image_files:
        image_path = os.path.join(input_dir, filename)

        # Horizontale Linien im Bild erkennen
        horizontal_lines = detect_horizontal_lines(image_path)

        # Text zwischen den erkannten Linien extrahieren
        texts = extract_text_between_lines(image_path, horizontal_lines)

        # Überprüfen, ob Texte extrahiert wurden
        if texts:
            # Extrahierte Texte in separate Listen entpacken
            ata, part_no, serial_no = zip(*texts)

            ataChapter += (list(ata))
            partNumber += (list(part_no))
            serialNumber += (list(serial_no))
        else:
            print(f"Keine Texte in {filename} extrahiert.")

    # Daten in ein Dictionary packen
    data = {
        'ATA': ataChapter,
        'PART NO.': partNumber,
        'SERIAL NO.': serialNumber
    }

    # Daten in ein DataFrame umwandeln
    df = pd.DataFrame(data)

    # DataFrame in eine Excel-Datei speichern
    df.to_excel(output_file, index=False)

    print("Daten erfolgreich in Excel gespeichert.")

# Funktion anwenden, um alle Bilder zu verarbeiten und Ergebnisse zu speichern
extract_texts_to_excel_google(output_dir_cropped, os.path.join(output_dir_results, 'Ergebnis_mit_GoogleVisionAPI.xlsx'))

end_time = time.time()
elapsed_time = end_time - start_time
print(f"Die Ausführung dauerte {elapsed_time:.6f} Sekunden")

# Modul 10: Google Vision API Ergebnisdatei formatieren

# Excel-Datei laden
file_path = os.path.join(output_dir_results, 'Ergebnis_mit_GoogleVisionAPI.xlsx')
df = pd.read_excel(file_path)

# Funktion, um Werte in der 'ATA'-Spalte von Liste zu String zu konvertieren
def convert_first_column_to_string(cell):
    if isinstance(cell, list):
        return ' '.join(cell)
    elif isinstance(cell, str) and cell.startswith("[") and cell.endswith("]"):
        return cell.strip("[]").strip("'")
    return cell

# Funktion auf die 'ATA'-Spalte anwenden
df['ATA'] = df['ATA'].apply(convert_first_column_to_string)

# Modifiziertes DataFrame in die gleiche Excel-Datei speichern
df.to_excel(file_path, index=False)

# Modul 11: Vergleichsdateien (Referenzdaten+OCR Ergebnisse) für Qualitätstest erstellen

# Dateipaare definieren, die zusammengeführt werden sollen
file_pairs = [
    (referenzdatei,
     os.path.join(output_dir_results, '2.Ergebnis_mit_Pytesseract.xlsx')),
    (referenzdatei,
     os.path.join(output_dir_results, '1.Ergebnis_mit_Pytesseract.xlsx')),
    (referenzdatei,
     os.path.join(output_dir_results, 'Ergebnis_mit_GoogleVisionAPI.xlsx'))
]

# Jedes Paar verarbeiten, um zusammengeführte Dateien zu erstellen
for original_file, ocr_file in file_pairs:
    # Daten aus beiden Dateien laden, erste Zeile überspringen
    original_df = pd.read_excel(original_file, skiprows=1)
    ocr_df = pd.read_excel(ocr_file, skiprows=1)

    # Sicherstellen, dass beide DataFrames die gleichen Spalten haben
    if len(original_df.columns) != 3 or len(ocr_df.columns) != 3:
        print(f"Fehler: Eine der Dateien ({original_file}, {ocr_file}) hat nicht genau 3 Spalten.")
        continue

    # Spalten umbenennen für Klarheit
    merged_data = {
        'ataChapter': original_df.iloc[:, 0],
        'partNumber': original_df.iloc[:, 1],
        'serialNumber': original_df.iloc[:, 2],
        'ataChapter_OCR': ocr_df.iloc[:, 0],
        'partNumber_OCR': ocr_df.iloc[:, 1],
        'serialNumber_OCR': ocr_df.iloc[:, 2],
    }

    # Zusammengeführtes DataFrame erstellen
    merged_df = pd.DataFrame(merged_data)

    # Ausgabedateinamen erstellen
    output_filename = os.path.join(output_dir_results, f'Qualitätstestdatei_{os.path.basename(ocr_file)}')
    merged_df.to_excel(output_filename, index=False)

print("Dateien wurden erfolgreich zusammengeführt.")

# Modul 12: Qualitätstest mit den in Exceldateien extrahierten Daten durchführen

# Dateien definieren, die verarbeitet werden sollen
file_paths = [
    os.path.join(output_dir_results, 'Qualitätstestdatei_1.Ergebnis_mit_Pytesseract.xlsx'),
    os.path.join(output_dir_results, 'Qualitätstestdatei_2.Ergebnis_mit_Pytesseract.xlsx'),
    os.path.join(output_dir_results, 'Qualitätstestdatei_Ergebnis_mit_GoogleVisionAPI.xlsx')
]

# Listen für Qualitäten und WER initialisieren
accuracies = []
wers = []

# Jede Datei verarbeiten
for file_path in file_paths:
    # Daten laden
    df = pd.read_excel(file_path)

    # Alle relevanten Spalten in String-Format konvertieren
    for col in ['ataChapter', 'partNumber', 'serialNumber', 'ataChapter_OCR', 'partNumber_OCR', 'serialNumber_OCR']:
        df[col] = df[col].astype(str)

    # Zähler initialisieren
    total_cells = 0
    mismatch_count = 0

    # Spalten vergleichen und Abweichungen zählen
    for col in ['ataChapter', 'partNumber', 'serialNumber']: 
        ocr_col = col + '_OCR'
        mismatches = df[col].str.strip() != df[ocr_col].str.strip()
        mismatch_count += mismatches.sum()
        total_cells += len(df[col])

    # Qualität berechnen
    accuracy = (total_cells - mismatch_count) / total_cells
    accuracies.append(accuracy)

    # WER (Word Error Rate) berechnen und zur Liste hinzufügen
    wer = (1 - accuracy) * 100
    wers.append(wer)

    # Arbeitsmappe und Blatt laden
    wb = load_workbook(file_path)
    sheet = wb.active

    # Schriftfarbe für abweichende Zellen anpassen
    for index, row in df.iterrows():
        for col_idx, col in enumerate(['ataChapter', 'partNumber', 'serialNumber']):
            original_cell = sheet.cell(row=index + 2, column=col_idx + 1)  # Anpassung wegen Header und 0-Index
            ocr_cell = sheet.cell(row=index + 2, column=col_idx + 4)
            if str(original_cell.value).strip() != str(ocr_cell.value).strip():
                ocr_cell.font = Font(color="FF0000")

    # Arbeitsmappe speichern
    wb.save(file_path)

# Genauigkeiten und WER ausgeben
for i, (accuracy, wer) in enumerate(zip(accuracies, wers)):
    print(f"Genauigkeit der OCR-Daten in {file_paths[i]}: {accuracy*100:.2f}%")
    print(f"Word Error Rate (WER) der OCR-Daten in {file_paths[i]}: {wer:.2f}%")
